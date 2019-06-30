import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os

os.chdir('C:\Git\Python\MoshHamedaniCourse\ExcelSpreadsheets')




def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                min_col=4,
                min_row=2, 
                max_row=sheet.max_row,
                max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'F2')

    workbook.save(filename)

filename = input("> ")
process_workbook(filename)